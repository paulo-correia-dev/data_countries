import logging
import json
import sys

import openpyxl
import requests
from openpyxl.styles import Alignment, Font

logging.basicConfig(filename='program.log', level=logging.INFO, format="%(asctime)s \ %(levelname)s \ %(filename)s \ %(lineno)i \ %(message)s")

try:
    url = 'https://restcountries.com/v2/all'
    response = requests.get(url=url)
    response = json.loads(response.content)
    logging.info('Conexão à API estabelecida com sucesso.')
except:
    print('Conexão à API [NÃO] estabelecida.')
    logging.error('Conexão à API [NÃO] estabelecida.')
    sys.exit()

# logging.info('The program successfully made a get request to the API.')

try:
    book = openpyxl.Workbook()
    book.create_sheet('Pag 1')
    pag1 = book['Pag 1']
    del book['Sheet']

    logging.info('The program created the book and sheet of arquive xlsx.')

    pag1.append(['Countries List'])
    pag1.merge_cells('A1:D1')
    currentCell = pag1['A1']
    currentCell.alignment = Alignment(horizontal='center')
    fontStyle = Font(size="16", bold=True, color="4F4F4F")
    pag1.cell(row=1, column=1, value='Countries List').font = fontStyle

    pag1.append(["Name", "Capital", "Area", "Currencies"])
    fontStyle = Font(size="12", bold=True, color="808080")

    for rows in pag1.iter_rows(min_row=2, max_row=2):
        for e, cell in enumerate(rows):
            value_str_cell = str(cell.value)
            pag1.cell(row=2, column=e + 1, value=value_str_cell).font = fontStyle

    logging.info('The program defined the fields and formatting of the table.')

    a = 3
    for e, c in enumerate(response):
        var = response[e].get('capital', '-')
        if var == 'capital':
            var = c['capital']

        varArea = response[e].get('area', '-')
        if varArea == 'area':
            varArea = c['area']

        varCurrencies = response[e].get('currencies', '-')
        if varCurrencies != '-':
            varCurrencies = ''
            for q, d in enumerate(response[e]['currencies']):
                if varCurrencies == '':
                    varCurrencies += response[e]['currencies'][q]['code']
                else:
                    varCurrencies += ", " + response[e]['currencies'][q]['code']

        pag1.append([c['name'], var, varArea, varCurrencies])

        pag1['C' + str(a)].number_format = '##0,00.00'
        a += 1
except:
    logging.info('Houve um erro na criação do arquivo.')
    print('Houve um erro na criação do arquivo.')
    sys.exit()
# logging.info('The program handled the fields without information.')
# logging.info('The program stored the API results in xlsx archive.')

try:
    book.save('Countries List.xlsx')
    logging.info('File generated successfully.')
except:
    logging.error('Erro de gravação do arquivo.')
    print('Erro de gravação do arquivo.')
